from datetime import date, datetime, timedelta
from calendar import monthrange
from io import BytesIO
from typing import Any, List, Optional, Tuple

from fastapi import APIRouter, Depends, HTTPException, Request, Query
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from sqlalchemy.orm import Session, joinedload

from ..database import get_db
from .. import models
from ..templating import templates

router = APIRouter(prefix="/reports", tags=["reports"])

MONTH_NAMES_HE = (
    "ינואר",
    "פברואר",
    "מרץ",
    "אפריל",
    "מאי",
    "יוני",
    "יולי",
    "אוגוסט",
    "ספטמבר",
    "אוקטובר",
    "נובמבר",
    "דצמבר",
)


def _families_debt_report_data(
    db: Session, year: int, month: int
) -> tuple[dict, int]:
    """Families with negative balance (actual debt), including after partial «שולם».

    Lesson rows in the table are non-cancelled lessons in the selected month only
    (context); the card total is always the live family balance (amount owed).
    """
    last_day = monthrange(year, month)[1]
    start_d = date(year, month, 1)
    end_d = date(year, month, last_day)

    indebted = (
        db.query(models.Family)
        .options(joinedload(models.Family.students))
        .filter(models.Family.balance < 0)
        .order_by(models.Family.name)
        .all()
    )

    families: dict = {}
    grand_total = 0

    for fam in indebted:
        owed = -int(fam.balance or 0)
        if owed <= 0:
            continue
        grand_total += owed
        student_ids = [s.id for s in fam.students]
        student_names = sorted(
            {f"{s.first_name} {s.last_name}" for s in fam.students},
            key=lambda n: n,
        )
        parent_name = (fam.contact_name or "").strip()
        parent_phone = (fam.phone or "").strip()
        if not parent_name or not parent_phone:
            for s in fam.students:
                if not parent_name and (s.parent_name or "").strip():
                    parent_name = (s.parent_name or "").strip()
                if not parent_phone and (s.parent_phone or "").strip():
                    parent_phone = (s.parent_phone or "").strip()
                if parent_name and parent_phone:
                    break

        lessons_in_month: list = []
        if student_ids:
            lessons_in_month = (
                db.query(models.Lesson)
                .options(joinedload(models.Lesson.student))
                .filter(models.Lesson.student_id.in_(student_ids))
                .filter(models.Lesson.lesson_date >= start_d)
                .filter(models.Lesson.lesson_date <= end_d)
                .filter(models.Lesson.status != "cancelled")
                .order_by(models.Lesson.lesson_date, models.Lesson.start_time)
                .all()
            )

        key = f"fam_{fam.id}"
        families[key] = {
            "student_names": student_names,
            "parent_name": parent_name,
            "parent_phone": parent_phone,
            "lessons": lessons_in_month,
            "total": owed,
            "family_balance": int(fam.balance or 0),
        }

    return families, grand_total


def _lesson_status_label(lesson: models.Lesson) -> str:
    if lesson.status == "completed":
        return "הושלם"
    return "מתוכנן"


def _month_year_from_query(month_param: Optional[str], today: date) -> Tuple[int, int]:
    """Parse ?month=YYYY-MM or fall back to current calendar month."""
    if month_param and len(month_param) >= 7:
        try:
            y = int(month_param[0:4])
            m = int(month_param[5:7])
            if 1 <= m <= 12 and 2000 <= y <= 2100:
                return y, m
        except ValueError:
            pass
    return today.year, today.month


def _shift_month(year: int, month: int, delta: int) -> Tuple[int, int]:
    m = month + delta
    y = year
    while m > 12:
        m -= 12
        y += 1
    while m < 1:
        m += 12
        y -= 1
    return y, m


def _month_key(year: int, month: int) -> str:
    return f"{year}-{month:02d}"


def _last_day_of_month(year: int, month: int) -> date:
    return date(year, month, monthrange(year, month)[1])


def _lesson_in_monthly_report(lesson: models.Lesson) -> bool:
    """Lessons that count in monthly financial / teaching stats (not pending attendance)."""
    if lesson.status == "cancelled":
        return False
    att = (lesson.attendance or "expected").lower()
    return att in ("arrived", "no_show")


def _family_balance_on_or_before(db: Session, family_id: int, on_or_before: date) -> int:
    tx = (
        db.query(models.BalanceTransaction)
        .filter(models.BalanceTransaction.family_id == family_id)
        .filter(models.BalanceTransaction.txn_date <= on_or_before)
        .order_by(
            models.BalanceTransaction.txn_date.desc(),
            models.BalanceTransaction.id.desc(),
        )
        .first()
    )
    if tx:
        return int(tx.balance_after)
    return 0


def _lesson_net_aligned_with_calendar_balance(lesson: models.Lesson) -> int:
    """Cash effect on balance — must stay aligned with lessons._lesson_payment_net_for_balance."""
    if lesson.status == "cancelled":
        return 0
    c = int(lesson.price or 0)
    if lesson.is_paid:
        p = int(lesson.paid_amount) if lesson.paid_amount is not None else c
        return p - c
    if not bool(getattr(lesson, "payment_finalized", False)):
        return 0
    att = (getattr(lesson, "attendance", None) or "expected").lower()
    if att == "no_show":
        return 0
    return -c


def _family_balance_from_lessons_up_to(db: Session, family_id: int, through: date) -> int:
    """Net family balance from all non-cancelled lessons on or before ``through`` (calendar / ledger rules)."""
    sids = [
        s.id
        for s in db.query(models.Student)
        .filter(models.Student.family_id == family_id)
        .all()
    ]
    if not sids:
        return 0
    lessons = (
        db.query(models.Lesson)
        .filter(models.Lesson.student_id.in_(sids))
        .filter(models.Lesson.status != "cancelled")
        .filter(models.Lesson.lesson_date <= through)
        .order_by(
            models.Lesson.lesson_date,
            models.Lesson.start_time,
            models.Lesson.id,
        )
        .all()
    )
    if not lessons:
        return 0
    first_d = min(l.lesson_date for l in lessons)
    bal = _family_balance_on_or_before(db, family_id, first_d - timedelta(days=1))
    for L in lessons:
        bal += _lesson_net_aligned_with_calendar_balance(L)
    return bal


def _family_first_reportable_month_bounds(
    db: Session, family_id: int
) -> Optional[Tuple[int, int]]:
    """First calendar month that has a lesson counted in the monthly report (הגיע/לא הגיע)."""
    sids = [
        s.id
        for s in db.query(models.Student)
        .filter(models.Student.family_id == family_id)
        .all()
    ]
    if not sids:
        return None
    lessons = (
        db.query(models.Lesson)
        .filter(models.Lesson.student_id.in_(sids))
        .filter(models.Lesson.status != "cancelled")
        .all()
    )
    best: Optional[date] = None
    for L in lessons:
        if not _lesson_in_monthly_report(L):
            continue
        d = L.lesson_date
        if best is None or d < best:
            best = d
    if best is None:
        return None
    return best.year, best.month


def _family_opening_carry_at_month_start(
    db: Session, family_id: int, year: int, month: int
) -> int:
    """יתרה בתחילת החודש: סגירה שרשרתית מהחודש הראשון עם שיעור בדוח, לא יומן גולמי בלבד."""
    fm = _family_first_reportable_month_bounds(db, family_id)
    if fm is None:
        py, pm = _shift_month(year, month, -1)
        return _family_balance_from_lessons_up_to(db, family_id, _last_day_of_month(py, pm))
    fy, fm_m = fm
    if (year, month) < (fy, fm_m):
        py, pm = _shift_month(year, month, -1)
        return _family_balance_from_lessons_up_to(db, family_id, _last_day_of_month(py, pm))
    py, pm = _shift_month(fy, fm_m, -1)
    carry = _family_balance_on_or_before(db, family_id, _last_day_of_month(py, pm))
    y, m = fy, fm_m
    while (y, m) < (year, month):
        ch, pd = _family_reportable_lesson_totals(db, family_id, y, m)
        carry = carry + pd - ch
        y, m = _shift_month(y, m, 1)
    return carry


def _family_month_snapshot(db: Session, family_id: int, year: int, month: int) -> dict[str, Any]:
    """חיוב/שולם מהשיעורים. יתרה סוף חודש = מעבר + שולם − חיוב; המעבר הוא סוף חודש קודם באותה שרשרת (מתאים למעקב חוב)."""
    ch, pd = _family_reportable_lesson_totals(db, family_id, year, month)
    carry = _family_opening_carry_at_month_start(db, family_id, year, month)
    end_bal = carry + pd - ch
    return {
        "year": year,
        "month": month,
        "month_key": _month_key(year, month),
        "month_label_he": f"{MONTH_NAMES_HE[month - 1]} {year}",
        "carry_over": carry,
        "month_charge": ch,
        "month_paid": pd,
        "end_balance": end_bal,
    }


def _chain_reset_during_month(db: Session, family_id: int, year: int, month: int) -> bool:
    """True if running balance reached ≥0 at some point during this calendar month."""
    py, pm = _shift_month(year, month, -1)
    prev_end = _last_day_of_month(py, pm)
    carry = _family_balance_on_or_before(db, family_id, prev_end)
    mk = _month_key(year, month)
    txs = (
        db.query(models.BalanceTransaction)
        .filter(models.BalanceTransaction.family_id == family_id)
        .filter(models.BalanceTransaction.month_key == mk)
        .order_by(
            models.BalanceTransaction.txn_date.asc(),
            models.BalanceTransaction.id.asc(),
        )
        .all()
    )
    running = carry
    for tx in txs:
        running += int(tx.balance_after) - int(tx.balance_before)
        if running >= 0:
            return True
    return running >= 0


def _family_should_show_debt_trail(
    carry_over: int, end_balance: int, debt_trail: List[dict[str, Any]]
) -> bool:
    """Breadcrumb only while some step still reflects חוב (negative carry or end). When fully settled — hide."""

    def _has_debt(a: int, b: int) -> bool:
        return int(a) < 0 or int(b) < 0

    if not debt_trail:
        return False
    if _has_debt(carry_over, end_balance):
        return True
    return any(
        _has_debt(s.get("carry_over", 0), s.get("end_balance", 0)) for s in debt_trail
    )


def _build_debt_trail(db: Session, family_id: int, year: int, month: int) -> List[dict[str, Any]]:
    """Newest month last in list (RTL display can reverse). We store oldest → newest."""
    trail: List[dict[str, Any]] = []
    cy, cm = year, month
    for _ in range(60):
        snap = _family_month_snapshot(db, family_id, cy, cm)
        trail.insert(0, snap)
        if snap["carry_over"] == 0:
            break
        py, pm = _shift_month(cy, cm, -1)
        if _chain_reset_during_month(db, family_id, py, pm):
            prev_snap = _family_month_snapshot(db, family_id, py, pm)
            trail.insert(0, prev_snap)
            break
        cy, cm = py, pm
    return trail


def _payment_method_label_he(code: Optional[str]) -> str:
    c = str(code or "").lower().strip()
    if c == "cash":
        return "מזומן"
    if c == "bit":
        return "ביט"
    if c == "paybox":
        return "פייבוקס"
    if c == "other":
        return "אחר"
    return ""


def _lesson_type_label_he(lesson: models.Lesson) -> str:
    if getattr(lesson, "is_group_lesson", False):
        return "קבוצתי"
    return "פרטי"


def _lesson_paid_display(lesson: models.Lesson) -> int:
    if lesson.is_paid:
        if lesson.paid_amount is not None:
            return int(lesson.paid_amount)
        return int(lesson.price or 0)
    return 0


def _family_reportable_lesson_totals(
    db: Session, family_id: int, year: int, month: int
) -> Tuple[int, int]:
    """חיוב/שולם לחודש לפי שיעורים בפועל — לא סכימת שורות יומן (כל עדכון הוסיף שורה מלאה)."""
    start_d = date(year, month, 1)
    end_d = _last_day_of_month(year, month)
    sids = [
        s.id
        for s in db.query(models.Student)
        .filter(models.Student.family_id == family_id)
        .all()
    ]
    if not sids:
        return 0, 0
    lessons = (
        db.query(models.Lesson)
        .filter(models.Lesson.student_id.in_(sids))
        .filter(models.Lesson.lesson_date >= start_d)
        .filter(models.Lesson.lesson_date <= end_d)
        .filter(models.Lesson.status != "cancelled")
        .all()
    )
    ch = 0
    pd = 0
    for L in lessons:
        if not _lesson_in_monthly_report(L):
            continue
        ch += int(L.price or 0)
        pd += _lesson_paid_display(L)
    return ch, pd


def _lesson_row_class(lesson: models.Lesson) -> str:
    charge = int(lesson.price or 0)
    fin = bool(getattr(lesson, "payment_finalized", False))
    if not fin and not lesson.is_paid:
        return "pending"
    paid = _lesson_paid_display(lesson)
    if not lesson.is_paid:
        return "unpaid"
    if paid < charge:
        return "partial"
    return "paid"


def _build_monthly_report_context(db: Session, year: int, month: int) -> dict[str, Any]:
    last_d = _last_day_of_month(year, month)
    start_d = date(year, month, 1)

    # All families that have students (optionally filter later)
    all_families = (
        db.query(models.Family)
        .options(joinedload(models.Family.students))
        .order_by(models.Family.name)
        .all()
    )

    report_rows: List[dict[str, Any]] = []
    total_month_charge = 0
    total_month_paid = 0
    total_end = 0
    lessons_done_count = 0
    families_with_lessons: set[int] = set()

    for fam in all_families:
        if not fam.students:
            continue
        student_ids = [s.id for s in fam.students]
        lessons_m = (
            db.query(models.Lesson)
            .options(joinedload(models.Lesson.student))
            .filter(models.Lesson.student_id.in_(student_ids))
            .filter(models.Lesson.lesson_date >= start_d)
            .filter(models.Lesson.lesson_date <= last_d)
            .filter(models.Lesson.status != "cancelled")
            .order_by(models.Lesson.lesson_date, models.Lesson.start_time)
            .all()
        )
        report_lessons = [x for x in lessons_m if _lesson_in_monthly_report(x)]
        snap = _family_month_snapshot(db, fam.id, year, month)

        has_activity = len(report_lessons) > 0
        has_balance_thread = snap["carry_over"] != 0 or snap["end_balance"] != 0
        if not has_activity and not has_balance_thread:
            continue

        if has_activity:
            families_with_lessons.add(fam.id)

        total_month_charge += snap["month_charge"]
        total_month_paid += snap["month_paid"]
        total_end += snap["end_balance"]

        for _ in report_lessons:
            lessons_done_count += 1

        contact_name = (fam.contact_name or "").strip()
        phone = (fam.phone or "").strip()
        if not contact_name or not phone:
            for s in fam.students:
                if not contact_name and (s.parent_name or "").strip():
                    contact_name = (s.parent_name or "").strip()
                if not phone and (s.parent_phone or "").strip():
                    phone = (s.parent_phone or "").strip()
                if contact_name and phone:
                    break

        end_b = snap["end_balance"]
        if end_b < 0:
            filt = "debt"
        elif end_b > 0:
            filt = "credit"
        else:
            filt = "balanced"

        students_out: List[dict[str, Any]] = []
        for st in sorted(fam.students, key=lambda x: (x.last_name, x.first_name)):
            st_lessons = [L for L in report_lessons if L.student_id == st.id]
            ch_sum = sum(int(L.price or 0) for L in st_lessons)
            pd_sum = sum(_lesson_paid_display(L) for L in st_lessons)
            lesson_rows = []
            for L in st_lessons:
                lesson_rows.append(
                    {
                        "lesson_date": L.lesson_date,
                        "start_time": L.start_time,
                        "end_time": L.end_time,
                        "price": int(L.price or 0),
                        "paid_display": _lesson_paid_display(L),
                        "is_paid": bool(L.is_paid),
                        "payment_method": (L.payment_method or "").strip(),
                        "type_he": _lesson_type_label_he(L),
                        "method_he": _payment_method_label_he(L.payment_method)
                        if L.is_paid
                        else "",
                        "row_class": _lesson_row_class(L),
                    }
                )
            students_out.append(
                {
                    "id": st.id,
                    "name": f"{st.first_name} {st.last_name}",
                    "charge_sum": ch_sum,
                    "paid_sum": pd_sum,
                    "lesson_count": len(st_lessons),
                    "lessons": lesson_rows,
                }
            )

        debt_trail = _build_debt_trail(db, fam.id, year, month)
        show_trail = _family_should_show_debt_trail(
            snap["carry_over"], snap["end_balance"], debt_trail
        )

        report_rows.append(
            {
                "id": fam.id,
                "name": fam.name or "משפחה",
                "contact_name": contact_name,
                "phone": phone,
                "carry_over": snap["carry_over"],
                "month_charge": snap["month_charge"],
                "month_paid": snap["month_paid"],
                "end_balance": snap["end_balance"],
                "lesson_count": len(report_lessons),
                "student_count": len(fam.students),
                "filter_class": filt,
                "debt_trail": debt_trail,
                "show_debt_trail": show_trail,
                "students": students_out,
            }
        )

    n_debt = sum(1 for r in report_rows if r["filter_class"] == "debt")
    n_credit = sum(1 for r in report_rows if r["filter_class"] == "credit")
    n_balanced = sum(1 for r in report_rows if r["filter_class"] == "balanced")
    n_all = len(report_rows)

    return {
        "monthly_families": report_rows,
        "summary": {
            "lessons_done": lessons_done_count,
            "families_with_lessons": len(families_with_lessons),
            "month_charge": total_month_charge,
            "month_paid": total_month_paid,
            "end_balance_sum": total_end,
            "families_in_debt_end": n_debt,
        },
        "filter_counts": {
            "all": n_all,
            "debt": n_debt,
            "credit": n_credit,
            "balanced": n_balanced,
        },
    }


def _lesson_duration_hours(lesson: models.Lesson) -> float:
    start_dt = datetime.combine(lesson.lesson_date, lesson.start_time)
    end_dt = datetime.combine(lesson.lesson_date, lesson.end_time)
    if end_dt <= start_dt:
        end_dt += timedelta(days=1)
    return max(0.0, (end_dt - start_dt).total_seconds() / 3600.0)


def _monthly_teaching_summary(db: Session, year: int, month: int) -> Tuple[float, int]:
    """Total scheduled teaching hours in month (non-cancelled) and lesson count."""
    last_day = monthrange(year, month)[1]
    start_d = date(year, month, 1)
    end_d = date(year, month, last_day)
    lessons = (
        db.query(models.Lesson)
        .filter(models.Lesson.lesson_date >= start_d)
        .filter(models.Lesson.lesson_date <= end_d)
        .filter(models.Lesson.status != "cancelled")
        .all()
    )
    total_h = sum(_lesson_duration_hours(l) for l in lessons)
    return total_h, len(lessons)


def _format_hours_display(hours: float) -> str:
    if abs(hours - round(hours)) < 0.05:
        return str(int(round(hours)))
    s = f"{hours:.1f}"
    return s.rstrip("0").rstrip(".")


def _load_openpyxl():
    """Import only when exporting Excel so the app starts even if the image lacks openpyxl (until rebuild)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        return Workbook, Alignment, Font, PatternFill
    except ImportError as exc:
        raise HTTPException(
            status_code=503,
            detail=(
                "חסרה חבילת openpyxl בקונטיינר. "
                "עצרי את השרת והריצי מתיקיית הפרויקט: "
                "docker compose build --no-cache app && docker compose up -d"
            ),
        ) from exc


@router.get("/", response_class=HTMLResponse)
def reports_page(
    request: Request,
    month: Optional[str] = Query(None, description="YYYY-MM"),
    db: Session = Depends(get_db),
):
    today = date.today()
    year, mon = _month_year_from_query(month, today)
    month_input = f"{year}-{mon:02d}"
    py, pm = _shift_month(year, mon, -1)
    ny, nm = _shift_month(year, mon, 1)
    report_month_label = f"{MONTH_NAMES_HE[mon - 1]} {year}"

    ctx = _build_monthly_report_context(db, year, mon)
    return templates.TemplateResponse(
        "monthly_report.html",
        {
            "request": request,
            "month_input": month_input,
            "report_month_label": report_month_label,
            "prev_month_q": f"{py}-{pm:02d}",
            "next_month_q": f"{ny}-{nm:02d}",
            **ctx,
        },
    )


@router.get("/api/monthly-data")
def reports_monthly_data_api(
    month: Optional[str] = Query(None, description="YYYY-MM"),
    db: Session = Depends(get_db),
):
    """JSON snapshot for דוח חודשי — refresh after payments without full reload."""
    today = date.today()
    y, m = _month_year_from_query(month, today)
    ctx = _build_monthly_report_context(db, y, m)
    families_out: List[dict[str, Any]] = []
    for f in ctx["monthly_families"]:
        families_out.append(
            {
                "id": f["id"],
                "filter_class": f["filter_class"],
                "end_balance": f["end_balance"],
                "carry_over": f["carry_over"],
                "month_charge": f["month_charge"],
                "month_paid": f["month_paid"],
                "show_debt_trail": f["show_debt_trail"],
                "debt_trail": [
                    {
                        "month_key": s["month_key"],
                        "month_label_he": s["month_label_he"],
                        "carry_over": s["carry_over"],
                        "end_balance": s["end_balance"],
                    }
                    for s in f["debt_trail"]
                ],
                "students": [
                    {
                        "id": s["id"],
                        "name": s["name"],
                        "charge_sum": s["charge_sum"],
                        "paid_sum": s["paid_sum"],
                        "lesson_count": s["lesson_count"],
                    }
                    for s in f["students"]
                ],
            }
        )
    return JSONResponse(
        content={
            "month_input": f"{y}-{m:02d}",
            "summary": ctx["summary"],
            "filter_counts": ctx["filter_counts"],
            "families": families_out,
        }
    )


@router.get("/export.xlsx")
def reports_export_excel(
    month: Optional[str] = Query(None, description="YYYY-MM (same as דוח; default נוכחי)"),
    db: Session = Depends(get_db),
):
    """Export family-balance debt report to Excel (Hebrew, RTL sheet)."""
    Workbook, Alignment, Font, PatternFill = _load_openpyxl()

    today = date.today()
    ey, em = _month_year_from_query(month, today)
    families, grand_total = _families_debt_report_data(db, ey, em)

    wb = Workbook()
    ws = wb.active
    ws.title = "חובות"

    try:
        ws.sheet_view.rightToLeft = True
    except Exception:
        pass

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    hdr_fill = PatternFill("solid", fgColor="F1F5F9")

    r = 1
    c1 = ws.cell(
        row=r,
        column=1,
        value="דוח חובות (יתרת משפחה — כולל אחרי תשלום חלקי)",
    )
    c1.font = title_font
    c1.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    r += 1
    ws.cell(row=r, column=1, value=f"תאריך יצוא: {date.today().strftime('%d/%m/%Y')}")
    r += 1
    ws.cell(
        row=r,
        column=1,
        value=f"חודש שיעורים בטבלה: {MONTH_NAMES_HE[em - 1]} {ey}",
    )
    r += 1
    ws.cell(row=r, column=1, value=f'סה"כ לגבייה: {grand_total} ₪').font = Font(bold=True, size=12)
    r += 2

    if not families:
        ws.cell(row=r, column=1, value="אין חובות פתוחים.")
        r += 1
    else:
        for _key, data in families.items():
            ws.cell(row=r, column=1, value="תלמידים:").font = header_font
            ws.cell(row=r, column=2, value=" & ".join(data["student_names"]))
            r += 1
            if data["parent_name"]:
                ws.cell(row=r, column=1, value="הורה:")
                ws.cell(row=r, column=2, value=data["parent_name"])
                r += 1
            if data["parent_phone"]:
                ws.cell(row=r, column=1, value="טלפון:")
                ws.cell(row=r, column=2, value=data["parent_phone"])
                r += 1
            ws.cell(row=r, column=1, value=f'יתרת חוב משפחתית: {data["total"]} ₪').font = Font(
                bold=True
            )
            r += 1

            headers = ("תאריך", "תלמיד", "שעות", "חיוב (₪)", "שולם (₪)", "סטטוס")
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=r, column=col, value=h)
                cell.font = header_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="right", vertical="center")
            r += 1

            for lesson in data["lessons"]:
                s = lesson.student
                time_s = (
                    f"{lesson.start_time.strftime('%H:%M')} – {lesson.end_time.strftime('%H:%M')}"
                )
                paid_cell = ""
                if lesson.is_paid:
                    p = (
                        int(lesson.paid_amount)
                        if lesson.paid_amount is not None
                        else int(lesson.price or 0)
                    )
                    paid_cell = p
                else:
                    paid_cell = "—"
                ws.cell(row=r, column=1, value=lesson.lesson_date.strftime("%d/%m/%Y"))
                ws.cell(row=r, column=2, value=f"{s.first_name} {s.last_name}")
                ws.cell(row=r, column=3, value=time_s)
                ws.cell(row=r, column=4, value=lesson.price)
                ws.cell(row=r, column=5, value=paid_cell)
                ws.cell(row=r, column=6, value=_lesson_status_label(lesson))
                for col in range(1, 7):
                    ws.cell(row=r, column=col).alignment = Alignment(
                        horizontal="right", vertical="center"
                    )
                r += 1

            r += 1

    for col_letter, width in (
        ("A", 14),
        ("B", 22),
        ("C", 16),
        ("D", 12),
        ("E", 12),
        ("F", 12),
    ):
        ws.column_dimensions[col_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"unpaid-report-{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{fname}"',
        },
    )


@router.get("/export-monthly.xlsx")
def reports_export_monthly_xlsx(
    month: Optional[str] = Query(None, description="YYYY-MM"),
    db: Session = Depends(get_db),
):
    """Excel export aligned with the דוח חודשי page."""
    Workbook, Alignment, Font, PatternFill = _load_openpyxl()

    today = date.today()
    ey, em = _month_year_from_query(month, today)
    ctx = _build_monthly_report_context(db, ey, em)
    families = ctx["monthly_families"]
    summ = ctx["summary"]

    wb = Workbook()
    ws = wb.active
    ws.title = "דוח חודשי"

    try:
        ws.sheet_view.rightToLeft = True
    except Exception:
        pass

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    hdr_fill = PatternFill("solid", fgColor="F1F5F9")
    r = 1

    c1 = ws.cell(row=r, column=1, value="דוח חודשי — תשלומים ומשפחות")
    c1.font = title_font
    c1.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    r += 1
    ws.cell(
        row=r,
        column=1,
        value=f"חודש: {MONTH_NAMES_HE[em - 1]} {ey}",
    )
    r += 1
    ws.cell(row=r, column=1, value=f"תאריך יצוא: {date.today().strftime('%d/%m/%Y')}")
    r += 1
    ws.cell(
        row=r,
        column=1,
        value=(
            f"שיעורים בחודש: {summ['lessons_done']} · "
            f"שולם ‎₪{summ['month_paid']} מתוך חיוב ‎₪{summ['month_charge']}"
        ),
    )
    r += 2

    if not families:
        ws.cell(row=r, column=1, value="אין נתונים לחודש זה.")
    else:
        for fam in families:
            ws.cell(row=r, column=1, value=f"משפחה: {fam['name']}").font = header_font
            r += 1
            if fam["contact_name"]:
                ws.cell(row=r, column=1, value=f"איש קשר: {fam['contact_name']}")
                r += 1
            if fam["phone"]:
                ws.cell(row=r, column=1, value=f"טלפון: {fam['phone']}")
                r += 1
            ws.cell(
                row=r,
                column=1,
                value=(
                    f"חיוב חודש: {fam['month_charge']} · שולם: {fam['month_paid']} · "
                    f"יתרה סוף חודש: {fam['end_balance']}"
                ),
            ).font = Font(bold=True)
            r += 1

            for st in fam["students"]:
                if st["lesson_count"] == 0:
                    continue
                ws.cell(row=r, column=1, value=f"תלמיד: {st['name']}").font = header_font
                r += 1
                headers = ("תאריך", "שעות", "סוג", "אמצעי", "חיוב", "שולם")
                for col, h in enumerate(headers, start=1):
                    cell = ws.cell(row=r, column=col, value=h)
                    cell.font = header_font
                    cell.fill = hdr_fill
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                r += 1
                for row in st["lessons"]:
                    time_s = (
                        f"{row['start_time'].strftime('%H:%M')}–"
                        f"{row['end_time'].strftime('%H:%M')}"
                    )
                    method = row["method_he"] or ("לא שולם" if not row["is_paid"] else "")
                    ws.cell(row=r, column=1, value=row["lesson_date"].strftime("%d/%m/%Y"))
                    ws.cell(row=r, column=2, value=time_s)
                    ws.cell(row=r, column=3, value=row["type_he"])
                    ws.cell(row=r, column=4, value=method)
                    ws.cell(row=r, column=5, value=row["price"])
                    ws.cell(row=r, column=6, value=row["paid_display"])
                    for col in range(1, 7):
                        ws.cell(row=r, column=col).alignment = Alignment(
                            horizontal="right", vertical="center"
                        )
                    r += 1
                r += 1
            r += 1

    for col_letter, width in (("A", 14), ("B", 14), ("C", 12), ("D", 12), ("E", 10), ("F", 10)):
        ws.column_dimensions[col_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"monthly-report-{ey}-{em:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
